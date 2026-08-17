# -*- coding: utf-8 -*-
"""Generate Paymax_Test_Cases_Register.xlsx — master bilingual test-case register."""

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"d:\Ma3n\Automation Test For new Paymax\Paymax_Test_Cases_Register.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ALT_FILL = PatternFill("solid", fgColor="D6EAF8")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

HEADERS = [
    "TC ID",
    "Module (EN)",
    "Module (AR)",
    "Test Case Name (EN)",
    "Test Case Name (AR)",
    "Description (EN)",
    "Description (AR)",
    "Type",
    "Priority",
    "Automation Status",
    "Method Name",
    "Suite / Class",
    "Date Added",
    "Notes",
]

MODULE_ADM_EN = "Patient Admission / Medical File (Reception)"
MODULE_ADM_AR = "الملف الطبي / استقبال المريض (مكتب الدخول)"
MODULE_OPD_EN = "Outpatient Visits (OPD Clinics)"
MODULE_OPD_AR = "العيادات الخارجية"
SUITE_ADMISSION = "PatientAdmissionTests"
SUITE_ACTION_BAR = "PatientActionBarTests"
SUITE_OPD = "OutpatientCycleTests"
TODAY = date.today().isoformat()

CASES = [
    (
        "TC-ADM-001",
        "Happy path: create a patient with mandatory fields (unique 4-word Arabic name)",
        "المسار السعيد: إضافة مريض بالبيانات الإجبارية (اسم عربي رباعي فريد)",
        "Create a new patient using all mandatory fields with a unique 4-word Arabic name and verify successful save.",
        "إنشاء مريض جديد بكل الحقول الإجبارية باستخدام اسم عربي رباعي فريد والتحقق من نجاح الحفظ.",
        "Positive",
        "High",
        "Automated",
        "createPatientWithMandatoryFields",
    ),
    (
        "TC-ADM-002",
        "Negative: missing Arabic name must show an error message",
        "سلبي: الحفظ بدون الاسم العربي يظهر رسالة خطأ",
        "Attempt to save without Arabic name and verify that a validation error is displayed.",
        "محاولة الحفظ بدون إدخال الاسم العربي والتحقق من ظهور رسالة خطأ للتحقق.",
        "Negative",
        "High",
        "Automated",
        "savingWithoutMandatoryFieldsShowsValidation",
    ),
    (
        "TC-ADM-003",
        "Mobile field must completely reject alphabetic input",
        "حقل الموبايل يرفض الحروف الأبجدية بالكامل",
        "Enter alphabetic characters in the mobile field and verify they are rejected / not accepted.",
        "إدخال حروف أبجدية في حقل الموبايل والتحقق من رفضها بالكامل.",
        "Negative",
        "Medium",
        "Automated",
        "mobileFieldRejectsAlphabeticCharacters",
    ),
    (
        "TC-ADM-004",
        "Invalid Egyptian mobile (short / wrong prefix) must block save",
        "موبايل مصري غير صحيح (قصير / بادئة غلط) يمنع الحفظ",
        "Enter an invalid Egyptian mobile number and verify that save is blocked with an error.",
        "إدخال رقم موبايل مصري غير صحيح والتحقق من منع الحفظ وظهور خطأ.",
        "Negative",
        "High",
        "Automated",
        "invalidEgyptianMobileIsRejectedOnSave",
    ),
    (
        "TC-ADM-005",
        "Invalid National ID pattern must show a validation error on save",
        "الرقم القومي غير الصحيح يظهر خطأ تحقق عند الحفظ",
        "Enter an invalid National ID pattern and verify a validation error appears on save.",
        "إدخال رقم قومي غير مطابق للنمط الصحيح والتحقق من ظهور خطأ عند الحفظ.",
        "Negative",
        "High",
        "Automated",
        "invalidNationalIdIsRejectedOnSave",
    ),
    (
        "TC-ADM-006",
        "Valid National ID must auto-populate age years/months",
        "الرقم القومي الصحيح يحسب السن تلقائيًا (سنين/شهور)",
        "Enter a valid National ID and verify that age years and months are auto-calculated.",
        "إدخال رقم قومي صحيح والتحقق من حساب السن تلقائيًا بالسنين والشهور.",
        "Positive",
        "High",
        "Automated",
        "validNationalIdAutoCalculatesAge",
    ),
    (
        "TC-ADM-007",
        "Future dates must be blocked by the birth-date picker (max=today)",
        "التواريخ المستقبلية ممنوعة في تاريخ الميلاد (أقصى تاريخ = اليوم)",
        "Verify that the birth-date picker blocks future dates (maximum selectable date is today).",
        "التحقق من أن منتقي تاريخ الميلاد يمنع اختيار تواريخ مستقبلية (أقصى تاريخ = اليوم).",
        "Negative",
        "Medium",
        "Automated",
        "futureBirthDatesAreDisabled",
    ),
    (
        "TC-ADM-008",
        "Picking a birth date manually must recalculate age fields",
        "اختيار تاريخ الميلاد يدويًا يعيد حساب حقول السن",
        "Select a birth date manually and verify that age fields are recalculated correctly.",
        "اختيار تاريخ ميلاد يدويًا والتحقق من إعادة حساب حقول السن بشكل صحيح.",
        "Positive",
        "Medium",
        "Automated",
        "manualBirthDateRecalculatesAge",
    ),
    (
        "TC-ADM-009",
        "Credit company (#@ اليكو) makes insurance number + expiry mandatory",
        "اختيار جهة الائتمان (#@ اليكو) يجعل رقم التأمين وتاريخ الانتهاء إجباريين",
        "Select credit company #@ اليكو and verify insurance number and expiry become mandatory.",
        "اختيار جهة الائتمان #@ اليكو والتحقق من أن رقم التأمين وتاريخ الانتهاء أصبحا إجباريين.",
        "Negative",
        "High",
        "Automated",
        "creditCompanyRequiresInsuranceNumberAndExpiry",
    ),
    (
        "TC-ADM-010",
        "New (جديد) must clear all inputs and reset dropdowns",
        "زرار «جديد» يمسح كل الحقول ويرجع القوائم المنسدلة للافتراضي",
        "Click New and verify that all inputs are cleared and dropdowns reset to default.",
        "الضغط على زرار جديد والتحقق من مسح كل الحقول وإعادة ضبط القوائم المنسدلة.",
        "Positive",
        "High",
        "Automated",
        "newButtonClearsInputsAndDropdowns",
    ),
    (
        "TC-ADM-011",
        "Search by patient code then open the matching result (temporary)",
        "البحث بكود المريض ثم فتح النتيجة المطابقة (مؤقت)",
        "Search by patient code, open the matching dropdown result, and verify patient data is populated. (Temporary workaround)",
        "البحث بكود المريض وفتح النتيجة من القائمة والتحقق من تعبئة بيانات المريض. (حل مؤقت)",
        "Positive",
        "High",
        "Automated",
        "searchPopulatesExistingPatientData",
    ),
    (
        "TC-ADM-012",
        "Created-by and created-date fields must be read-only",
        "حقلا «أُنشئ بواسطة» و«تاريخ الإنشاء» للقراءة فقط",
        "Verify that created-by and created-date fields are read-only.",
        "التحقق من أن حقلي أُنشئ بواسطة وتاريخ الإنشاء للقراءة فقط.",
        "Positive",
        "Medium",
        "Automated",
        "createdByAndCreatedDateAreReadOnly",
    ),
    (
        "TC-ADM-013",
        "Header date must match the actual current local date",
        "تاريخ الـ Header يطابق تاريخ اليوم الفعلي",
        "Verify that the header date displayed matches the current local system date.",
        "التحقق من أن التاريخ الظاهر في الـ Header يطابق تاريخ الجهاز الحالي.",
        "Positive",
        "Low",
        "Automated",
        "headerDateMatchesCurrentDate",
    ),
    (
        "TC-ADM-014",
        "TC8: Single-word Arabic name must show ERROR on Save",
        "TC8: اسم عربي من كلمة واحدة يظهر Error عند الحفظ",
        "Save a patient with a single-word Arabic name and verify an ERROR message is shown.",
        "حفظ مريض باسم عربي من كلمة واحدة والتحقق من ظهور رسالة Error.",
        "Negative",
        "High",
        "Automated",
        "singleWordNameShowsErrorOnSave",
    ),
    (
        "TC-ADM-015",
        "TC9: Two/three-word Arabic name shows WARNING + SUCCESS on Save",
        "TC9: اسم ثنائي/ثلاثي يظهر Warning + Success عند الحفظ (ويُقبل)",
        "Fill mandatory fields with a 2/3-word Arabic name, save, and verify both WARNING and SUCCESS appear (save is allowed).",
        "ملء الحقول الإجبارية باسم ثنائي/ثلاثي والحفظ والتحقق من ظهور Warning و Success معًا (الحفظ مسموح).",
        "Positive",
        "High",
        "Automated",
        "twoOrThreeWordNameShowsWarningOnBlur",
    ),
    (
        "TC-ADM-016",
        "TC10: Four-word Arabic name is accepted (Success)",
        "TC10: اسم رباعي يُقبل بنجاح (Success)",
        "Save a patient with a four-word Arabic name and verify successful save (Success).",
        "حفظ مريض باسم عربي رباعي والتحقق من نجاح الحفظ (Success).",
        "Positive",
        "High",
        "Automated",
        "quadrupleNameHappyPathShowsSuccessOnly",
    ),
    (
        "TC-ADM-017",
        "Full happy path: fill all mandatory + optional fields and save",
        "Happy Path كامل: ملء كل الحقول الإجبارية والاختيارية ثم الحفظ",
        "Fill all mandatory and optional fields then save and verify success.",
        "ملء كل الحقول الإجبارية والاختيارية ثم الحفظ والتحقق من النجاح.",
        "Positive",
        "High",
        "Automated",
        "fullHappyPathAllFieldsFilled",
    ),
    (
        "TC-ADM-018",
        "Full form reset: New clears all filled inputs and dropdowns",
        "إعادة ضبط كاملة: زرار «جديد» يمسح كل الحقول بعد الملء",
        "Fill the full form then click New and verify everything is cleared/reset.",
        "ملء الفورم بالكامل ثم الضغط على جديد والتحقق من مسح/إعادة ضبط كل شيء.",
        "Positive",
        "Medium",
        "Automated",
        "fullFormResetViaNewButton",
    ),
    (
        "TC-ADM-019",
        "Age years auto-calculate to 10 when DOB is exactly 10 years ago",
        "السن يُحسب تلقائيًا = 10 لما تاريخ الميلاد يبقى من 10 سنين بالظبط",
        "Set DOB to exactly 10 years ago and verify age years equals 10.",
        "ضبط تاريخ الميلاد ليكون قبل 10 سنوات بالضبط والتحقق أن السن = 10.",
        "Positive",
        "Medium",
        "Automated",
        "ageAutoCalculatesFromDateOfBirthTenYearsAgo",
    ),
    (
        "TC-ADM-020",
        "Credit company requires insurance number + expiry on Save (error when empty)",
        "جهة الائتمان تطلب رقم التأمين والانتهاء عند الحفظ (Error لو فاضيين)",
        "Select credit company without insurance number/expiry and verify error on save.",
        "اختيار جهة الائتمان بدون رقم التأمين/الانتهاء والتحقق من ظهور Error عند الحفظ.",
        "Negative",
        "High",
        "Automated",
        "creditCompanyMandatoryFieldsShowErrorWhenEmpty",
    ),
    (
        "TC-ADM-021",
        "Created-by and created-date are readonly and reject typing",
        "حقلا الإنشاء readonly وبيرفضوا الكتابة",
        "Attempt to type into created-by / created-date and verify the fields reject modification.",
        "محاولة الكتابة في حقلي الإنشاء والتحقق من رفض التعديل.",
        "Negative",
        "Medium",
        "Automated",
        "readOnlyFieldsRejectModification",
    ),
    (
        "TC-ADM-022",
        "Unknown patient modal: auto-generated patient code is read-only",
        "بوب أب المريض غير المعروف: كود المريض المولّد تلقائيًا للقراءة فقط",
        "Open the unknown-patient modal and verify the patient code input is displayed and readonly/disabled.",
        "فتح بوب أب المريض غير المعروف والتحقق من ظهور حقل كود المريض وأنه readonly أو disabled.",
        "Positive",
        "Medium",
        "Automated",
        "unknownPatientCodeIsReadOnly",
    ),
    (
        "TC-ADM-023",
        "Unknown patient modal: happy path save shows a success message",
        "بوب أب المريض غير المعروف: الحفظ الناجح يظهر رسالة Success",
        "Open the modal, set gender and visit date/time, save, and verify a SUCCESS message appears.",
        "فتح البوب أب وتحديد الجنس وتاريخ ووقت الزيارة والحفظ والتحقق من ظهور رسالة Success.",
        "Positive",
        "High",
        "Automated",
        "saveUnknownPatientHappyPath",
    ),
    (
        "TC-ADM-024",
        "Unknown patient modal: X button closes the modal",
        "بوب أب المريض غير المعروف: زرار X يقفل البوب أب",
        "Open the modal, click the close (X) button, and verify the modal is no longer visible.",
        "فتح البوب أب والضغط على زرار الإغلاق (X) والتحقق من اختفاء البوب أب.",
        "Positive",
        "Medium",
        "Automated",
        "closeUnknownPatientModal",
    ),
    (
        "TC-ADM-025",
        "Default الجهة value is $$نقدي 2019 on a fresh admission page",
        "القيمة الافتراضية للجهة هي $$نقدي 2019 عند فتح الصفحة",
        "Reload the admission page and assert الجهة defaults to $$نقدي 2019.",
        "إعادة تحميل صفحة الاستقبال والتحقق أن الجهة الافتراضية = $$نقدي 2019.",
        "Positive",
        "High",
        "Automated",
        "defaultClientDropdownValueIsCash2019",
    ),
    (
        "TC-ADM-026",
        "Identity type selection updates the adjacent ID-number field label",
        "اختيار إثبات الشخصية يحدّث عنوان حقل الرقم المجاور",
        "Select جواز سفر from إثبات الشخصية and assert the adjacent field label becomes جواز سفر.",
        "اختيار جواز سفر من إثبات الشخصية والتحقق أن عنوان الحقل المجاور يصبح جواز سفر.",
        "Positive",
        "High",
        "Automated",
        "identityTypeUpdatesAdjacentFieldLabel",
    ),
    (
        "TC-ADM-027",
        "نوع التعاقد defaults to العضو نفسه; DDL has العضو نفسه/مريض تابع; dependent defaults درجة القرابة",
        "نوع التعاقد افتراضيًا العضو نفسه؛ القائمة فيها العضو نفسه/مريض تابع؛ ومريض تابع يضبط درجة القرابة",
        "After credit جهة + شركة فرعية: assert نوع التعاقد enabled and defaults to العضو نفسه; open DDL and assert exactly العضو نفسه + مريض تابع; with العضو نفسه درجة القرابة stays disabled; select مريض تابع and assert درجة القرابة defaults to الزوج/الزوجة (readonly).",
        "بعد اختيار جهة آجل وشركة فرعية: التحقق أن نوع التعاقد مفعّل والافتراضي العضو نفسه؛ فتح القائمة والتحقق من وجود اختيارين فقط العضو نفسه ومريض تابع؛ مع العضو نفسه درجة القرابة تبقى معطّلة؛ اختيار مريض تابع والتحقق أن درجة القرابة الافتراضية الزوج/الزوجة (readonly).",
        "Positive",
        "High",
        "Automated",
        "cascadingContractAndDependentPatientRules",
    ),
    (
        "TC-ADM-028",
        "Top action bar: each active button redirects with dynamic patientCode",
        "شريط الإجراءات: كل زر نشط يوجّه برابط فيه كود المريض الديناميكي",
        "Open a patient profile and click دخول / عيادات خارجية / الأشعة / تحاليل / إسعاف وطوارئ; assert each URL matches the expected path and includes the dynamic patientCode (and clientId=52 where applicable).",
        "فتح بروفايل المريض والضغط على دخول / عيادات خارجية / الأشعة / تحاليل / إسعاف وطوارئ والتحقق أن كل رابط يطابق المسار المتوقع ويحتوي كود المريض الديناميكي (وclientId=52 عند اللزوم).",
        "Positive",
        "High",
        "Automated",
        "verifyDynamicActionBarRedirects",
    ),
    (
        "TC-ADM-029",
        "Top action bar: Patient Barcode opens popup with /print/barcode?id={code}",
        "شريط الإجراءات: باركود المريض يفتح نافذة منبثقة /print/barcode?id={code}",
        "Click باركود المريض, switch to the popup window, assert URL contains /print/barcode?id={patientCode}, close popup, and return to the main window (avoid getting stuck on native print dialog).",
        "الضغط على باركود المريض والتبديل للنافذة المنبثقة والتحقق أن الرابط يحتوي /print/barcode?id={كود المريض} ثم إغلاق النافذة والرجوع للنافذة الرئيسية دون التعطل على دايلوج الطباعة.",
        "Positive",
        "High",
        "Automated",
        "verifyPatientBarcodePopupWindow",
    ),
    (
        "TC-ADM-030",
        "Top action bar: under-development buttons stay inactive / do not break the app",
        "شريط الإجراءات: أزرار قيد التطوير تبقى غير نشطة ولا تكسر التطبيق",
        "With a patient open, verify قسطرة / غسيل كلى / المناظير / الأرشيف are inactive or leave the patient profile intact when clicked.",
        "مع فتح مريض، التحقق أن قسطرة / غسيل كلى / المناظير / الأرشيف غير نشطة أو تبقى على بروفايل المريض عند الضغط عليها.",
        "Negative",
        "Medium",
        "Automated",
        "verifyUnderDevelopmentActionButtonsDoNotBreakApp",
    ),
    (
        "TC-ADM-031",
        "External visit family (OPD/Rays/Lab): switching modes keeps the same visit code",
        "عائلة الزيارة الخارجية (عيادات/أشعة/تحاليل): التنقل بينهم يحافظ على نفس كود الزيارة",
        "For seeded patients already in OPD (30499) / Rays (30485) / Lab (30456), open the other external modes and assert كود الزيارة stays the same (mode changes in the URL only).",
        "لمرضى جاهزين في زيارة خارجي 30499 / أشعة 30485 / تحاليل 30456، فتح باقي أوضاع الزيارة الخارجية والتحقق أن كود الزيارة ثابت (يتغير الـmode في الرابط فقط).",
        "Positive",
        "High",
        "Automated",
        "verifyExternalVisitFamilyKeepsSameVisitCode",
    ),
    (
        "TC-ADM-032",
        "Incompatible active visit: switching visit type shows ERROR and stays on profile",
        "زيارة نشطة غير متوافقة: تغيير نوع الزيارة يظهر Error ويبقى على البروفايل",
        "For patients already in Inpatient (30496), Emergency (30495), or External family trying دخول/إسعاف: click an incompatible action and assert an ERROR toast and no navigation away from the patient profile.",
        "لمرضى في داخلي 30496 أو طوارئ 30495 أو زيارة خارجية يحاولون دخول/إسعاف: الضغط على إجراء غير متوافق والتحقق من ظهور رسالة Error وعدم مغادرة بروفايل المريض.",
        "Negative",
        "High",
        "Automated",
        "verifyIncompatibleVisitSwitchShowsError",
    ),
    (
        "TC-ADM-033",
        "Refresh on selected patient (/reception/patient/{code}) redirects to /reception",
        "عمل Refresh وصفحة مريض مختار (/reception/patient/{code}) يرجع إلى /reception",
        "Open/save a patient so URL is /reception/patient/{code}, refresh the browser, and assert the URL becomes exact /reception (not the patient detail URL).",
        "فتح/حفظ مريض بحيث يكون الرابط /reception/patient/{code}، عمل Refresh، والتحقق أن الرابط يصبح /reception بالظبط (مش صفحة تفاصيل المريض).",
        "Positive",
        "High",
        "Automated",
        "refreshOnPatientProfileRedirectsToReception",
    ),
    (
        "TC-ADM-034",
        "New (جديد) on selected patient navigates to exact /reception",
        "زرار جديد على مريض مختار يوجّه إلى /reception",
        "With a selected patient URL /reception/patient/{code}, click جديد and assert navigation to exact /reception.",
        "مع مريض مختار على /reception/patient/{code}، الضغط على جديد والتحقق من الانتقال إلى /reception بالظبط.",
        "Positive",
        "High",
        "Automated",
        "newButtonOnPatientProfileNavigatesToReception",
    ),
    (
        "TC-OPD-001",
        "Draft OPD visit: fill نوع الحالة + الشركة الفرعية, save → /services",
        "مسودة زيارة خارجية: ملء نوع الحالة والشركة الفرعية ثم الحفظ → /services",
        "Create/open a clean patient, click عيادات خارجية, assert draft URL (mode=opd), fill نوع الحالة + الشركة الفرعية, save, and assert URL becomes /clinic/visits/{visitCode}/services.",
        "إنشاء/فتح مريض نظيف، الضغط على عيادات خارجية، التحقق من رابط المسودة mode=opd، ملء نوع الحالة والشركة الفرعية والحفظ، والتحقق أن الرابط يصبح /clinic/visits/{كود الزيارة}/services.",
        "Positive",
        "High",
        "Automated",
        "createDraftVisitAndTransitionToServices",
    ),
    (
        "TC-OPD-002",
        "Active visit modal: continue same visit → services page",
        "بوب أب الزيارة المفتوحة: الإستكمال على نفس الزيارة → صفحة الخدمات",
        "Open patient 30499, click عيادات خارجية, assert open-visit modal, click patient-data-continue-same-visit-btn, assert /clinic/visits/{visitCode}/services.",
        "فتح مريض 30499، الضغط على عيادات خارجية، التحقق من بوب أب الزيارة المفتوحة، الضغط على الإستكمال على نفس الزيارة، والتحقق من /clinic/visits/{كود الزيارة}/services.",
        "Positive",
        "High",
        "Automated",
        "activeVisitModalContinueExisting",
    ),
    (
        "TC-OPD-003",
        "Close open visit blocked when unpaid — confirm نعم then تنبيه لم تُدفع بعد",
        "إغلاق الزيارة المفتوحة يُمنع لو الزيارة غير مدفوعة — تأكيد نعم ثم تنبيه",
        "Open unpaid patient 30552, click عيادات خارجية, click close-open-visit, confirm نعم on Swal, assert تنبيه message contains تعذر إغلاق الزيارة / لم تُدفع بعد, dismiss with حسناً, stay on patient profile.",
        "فتح مريض غير مدفوع 30552، عيادات خارجية، إغلاق الزيارة المفتوحة، تأكيد نعم، التحقق من تنبيه تعذر إغلاق الزيارة بسبب عدم الدفع، حسناً، والبقاء على بروفايل المريض.",
        "Negative",
        "High",
        "Automated",
        "closeVisitBlockedWhenVisitUnpaid",
    ),
    (
        "TC-OPD-004",
        "Add service inside saved visit and verify الإجمالي updates",
        "إضافة خدمة داخل زيارة محفوظة والتحقق من تحديث الإجمالي",
        "Inside a saved OPD visit, add a service and assert الإجمالي increases. (Deferred until business steps are clarified.)",
        "داخل زيارة خارجية محفوظة، إضافة خدمة والتحقق أن الإجمالي يزيد. (موقوف مؤقتًا لحين توضيح الخطوات.)",
        "Positive",
        "High",
        "Planned",
        "addServiceAndVerifyTotalUpdates",
    ),
    (
        "TC-ADM-035",
        "Reservation Agenda modal opens with current dates (from/to)",
        "نافذة أجندة الحجوزات تفتح بالتاريخ الحالي (من/إلى)",
        "Open Reservation Agenda modal and verify default dates (from/to) are set to today's date.",
        "فتح نافذة أجندة الحجوزات والتحقق من ضبط تاريخ البداية والنهاية افتراضيًا على تاريخ اليوم.",
        "Positive",
        "High",
        "Automated",
        "verifyReservationAgendaModalOpensWithCurrentDates",
    ),
    (
        "TC-ADM-036",
        "Reservation Agenda search by specific date returns matching results",
        "البحث في أجندة الحجوزات بتاريخ محدد يظهر النتيجة المطابقة",
        "Search reservations for specific date 2026-08-08 and verify exact row count in the grid.",
        "البحث في أجندة الحجوزات بتاريخ محدد (2026-08-08) والتحقق من ظهور العدد المطابق في الجدول.",
        "Positive",
        "High",
        "Automated",
        "verifyReservationSearchBySpecificDate",
    ),
    (
        "TC-OPD-005",
        "Services page header title and summary chips verification",
        "التحقق من عنوان الهيدر ومؤشرات ملخص صفحة الخدمات",
        "Verify patient name title and summary chips (total, paid, insurance %, remaining, services count) on services page.",
        "التحقق من اسم المريض في العنوان ومؤشرات الملخص (الإجمالي، المدفوع، التأمين، المتبقي، عدد الخدمات) في صفحة الخدمات.",
        "Positive",
        "High",
        "Automated",
        "verifyServicesPageHeaderAndSummaryChips",
    ),
    (
        "TC-OPD-006",
        "Services tabs active state and badge counts",
        "تبويبات صفحة الخدمات والشارات الرقمية",
        "Verify الخدمات tab is active by default and badge shows initial count (0) on fresh visit.",
        "التحقق أن تبويب الخدمات نشط افتراضيًا والشارة تظهر العدد الأولي (0) عند فتح زيارة جديدة.",
        "Positive",
        "Medium",
        "Automated",
        "verifyServicesTabsAndBadges",
    ),
    (
        "TC-OPD-007",
        "Add service form initial state and price field restrictions",
        "الحالة الأولية لنموذج إضافة خدمة وقيود أسعار الخدمات",
        "Verify إضافة للقائمة button disabled, prices read-only, favorites button visible, and empty message shown.",
        "التحقق من تعطيل زرار إضافة للقائمة، وأن أسعار الخدمة للقراءة فقط، وزرار المفضلة ظاهر ورسالة لا توجد خدمات تظهر.",
        "Positive",
        "Medium",
        "Automated",
        "addServiceFormInitialState",
    ),
    (
        "TC-OPD-008",
        "Add service to visit and save visit info",
        "إضافة خدمة للزيارة وحفظ بيانات الزيارة",
        "Select service/doctor/clinic, click add to list, save visit info, and verify successful save.",
        "اختيار الخدمة والطبيب والعيادة، الضغط على إضافة للقائمة، حفظ بيانات الزيارة والتحقق من الحفظ بنجاح.",
        "Positive",
        "High",
        "Automated",
        "addServiceUpdatesListAndTotal",
    ),
    (
        "TC-OPD-009",
        "Save visit info button shows success toast notification",
        "الضغط على حفظ بيانات الزيارة يظهر توست النجاح",
        "Click حفظ (save visit info) button and verify success toast notification appears.",
        "الضغط على زرار حفظ بيانات الزيارة والتحقق من ظهور رسالة النجاح (Toast).",
        "Positive",
        "High",
        "Automated",
        "saveVisitInfoShowsSuccessToast",
    ),
    (
        "TC-OPD-010",
        "OPD visit payment flow generates receipt and updates history table",
        "سداد خدمات الزيارة الخارجية ينشئ إيصالاً ويحدّث جدول الإيصالات",
        "Submit payment for visit services and verify a new receipt row appears in the payment history table.",
        "تنفيذ سداد خدمات الزيارة والتحقق من إضافة صف إيصال جديد في جدول سجل الإيصالات.",
        "Positive",
        "High",
        "Automated",
        "paymentCreatesReceiptAndUpdatesPaidChip",
    ),
    (
        "TC-OPD-011",
        "Services page back button navigates to previous page",
        "زرار العودة بصفحة الخدمات يرجع للصفحة السابقة",
        "Click opd-patient-services-back-btn button and verify navigation away from services page.",
        "الضغط على زرار العودة (رجوع) في صفحة الخدمات والتحقق من مغادرة الصفحة بنجاح.",
        "Positive",
        "Medium",
        "Automated",
        "backButtonNavigatesFromServicesPage",
    ),
    (
        "TC-OPD-012",
        "Service group filter limits available clinics",
        "فلترة مجموعة الخدمات تحدد العيادات المتاحة",
        "Select service group الرمد والليزك in opd-patient-services-service-group-select and verify clinic dropdown displays only عيادة 8.",
        "اختيار مجموعة الخدمات الرمد والليزك والتحقق أن قائمة العيادات المتاحة تحتوي فقط على عيادة 8.",
        "Positive",
        "High",
        "Automated",
        "serviceGroupFilterLimitsClinicOptions",
    ),
    (
        "TC-OPD-013",
        "ICD10 Diagnosis mandatory validation by entity (optional for النقدي2019, mandatory for اليكو)",
        "التحقق من إلزامية تشخيص ICD10 حسب الجهة (اختياري للنقدي2019 وإجباري لاليكو)",
        "Verify OPD visit with entity النقدي2019 saves without ICD10 diagnosis, while entity اليكو blocks save until ICD10 diagnosis is selected.",
        "التحقق أن الزيارة مع جهة النقدي2019 تحفظ بدون تشخيص ICD10، بينما تمنع جهة اليكو الحفظ إلا بعد اختيار تشخيص ICD10.",
        "Positive",
        "High",
        "Automated",
        "icdDiagnosisMandatoryValidationByEntity",
    ),
    (
        "TC-OPD-014",
        "Print barcode button opens correct barcode URL",
        "زرار طباعة الباركود يفتح الرابط الصحيح للباركود",
        "Click opd-patient-services-print-barcode-btn button and verify opened URL matches /clinic/visits/{visitCode}/barcode-print.",
        "الضغط على زرار طباعة الباركود والتحقق أن الرابط المفتوح يطابق /clinic/visits/{كود الزيارة}/barcode-print.",
        "Positive",
        "Medium",
        "Automated",
        "printBarcodeButtonOpensCorrectUrl",
    ),
    (
        "TC-OPD-015",
        "Claim form modal CRUD operations (add, download, delete claim file)",
        "عمليات نموذج المطالبة (إضافة، تحميل، حذف ملف المطالبة)",
        "Open claim modal via opd-patient-services-open-claim-btn, upload claim title and file, verify in grid, test download and delete buttons.",
        "فتح بوب أب المطالبة، رفع عنوان وملف المطالبة، التأكد من التواجد في الجدول، وتجربة أزرار التحميل والحذف.",
        "Positive",
        "Medium",
        "Automated",
        "claimModalAddDownloadAndDeleteOperations",
    ),
    (
        "TC-OPD-016",
        "Service short code lookup (valid code 4444 populates data, invalid code 91238128 triggers error message)",
        "التحقق من الكود المختصر للخدمة (الكود الصحيح 4444 يجيب البيانات، والكود الخاطئ 91238128 يظهر خطأ)",
        "Type short code 4444 + ENTER to verify service and price auto-populate. Type invalid code 91238128 + ENTER to verify error message كود الخدمة غير صحيح.",
        "إدخال الكود المختصر 4444 والضغط على ENTER والتحقق من تعبئة الخدمة والسعر تلقائيًا. إدخال الكود الخاطئ 91238128 والضغط على ENTER والتحقق من ظهور رسالة الخطأ كود الخدمة غير صحيح.",
        "Positive",
        "High",
        "Automated",
        "serviceShortCodeValidAndInvalidVerification",
    ),
    (
        "TC-OPD-017",
        "Approval number field state toggle (closed by default, enabled for service 444)",
        "حالة حقل رقم الموافقة (معطل افتراضيًا ويفتح للخدمة 444 التي تطلب موافقة)",
        "Verify approval number field is closed/disabled by default, and becomes enabled/open when service 444 is selected.",
        "التحقق أن حقل رقم الموافقة معطل/مغلق افتراضيًا للخدمات العادية، ويفتح للإدخال عند اختيار الخدمة 444.",
        "Positive",
        "High",
        "Automated",
        "approvalNumberFieldToggleVerification",
    ),
    (
        "TC-OPD-018",
        "Complete outpatient service draft and save cycle (group -> service 444 -> doctor -> clinic -> approval no -> draft list -> saved grid)",
        "السايكل الكاملة للخدمات الخارجية (مجموعة الخدمات -> الخدمة 444 -> الطبيب -> العيادة -> رقم الموافقة -> قائمة الانتظار -> الحفظ في الجريد)",
        "Execute full OPD service cycle: select service group, service 444, doctor, clinic, approval number 324, click add to list (Draft), then click save drafts and verify transfer to saved services grid.",
        "تنفيذ سايكل إضافة الخدمة بالكامل: اختيار مجموعة الخدمات والخدمة 444 والطبيب والعيادة ورقم الموافقة 324 والضغط على إضافة للقائمة لتنزل في خدمات في الانتظار ثم حفظ الخدمات والتحقق من انتقالها لجدول الخدمات المحفوظة.",
        "Positive",
        "High",
        "Automated",
        "completeServicesDraftAndSaveCycle",
    ),
]

# Suite override (default = PatientAdmissionTests)
SUITE_BY_ID = {
    "TC-ADM-028": SUITE_ACTION_BAR,
    "TC-ADM-029": SUITE_ACTION_BAR,
    "TC-ADM-030": SUITE_ACTION_BAR,
    "TC-ADM-031": SUITE_ACTION_BAR,
    "TC-ADM-032": SUITE_ACTION_BAR,
    "TC-OPD-001": SUITE_OPD,
    "TC-OPD-002": SUITE_OPD,
    "TC-OPD-003": SUITE_OPD,
    "TC-OPD-004": SUITE_OPD,
    "TC-OPD-005": SUITE_OPD,
    "TC-OPD-006": SUITE_OPD,
    "TC-OPD-007": SUITE_OPD,
    "TC-OPD-008": SUITE_OPD,
    "TC-OPD-009": SUITE_OPD,
    "TC-OPD-010": SUITE_OPD,
    "TC-OPD-011": SUITE_OPD,
    "TC-OPD-012": SUITE_OPD,
    "TC-OPD-013": SUITE_OPD,
    "TC-OPD-014": SUITE_OPD,
    "TC-OPD-015": SUITE_OPD,
    "TC-OPD-016": SUITE_OPD,
    "TC-OPD-017": SUITE_OPD,
    "TC-OPD-018": SUITE_OPD,
}

MODULE_BY_ID = {
    "TC-OPD-001": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-002": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-003": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-004": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-005": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-006": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-007": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-008": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-009": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-010": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-011": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-012": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-013": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-014": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-015": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-016": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-017": (MODULE_OPD_EN, MODULE_OPD_AR),
    "TC-OPD-018": (MODULE_OPD_EN, MODULE_OPD_AR),
}

DATE_BY_ID = {
    "TC-ADM-022": "2026-07-25",
    "TC-ADM-023": "2026-07-25",
    "TC-ADM-024": "2026-07-25",
    "TC-ADM-025": "2026-07-25",
    "TC-ADM-026": "2026-07-25",
    "TC-ADM-027": "2026-07-25",
    "TC-ADM-028": "2026-07-28",
    "TC-ADM-029": "2026-07-28",
    "TC-ADM-030": "2026-07-28",
    "TC-ADM-031": "2026-07-28",
    "TC-ADM-032": "2026-07-28",
    "TC-ADM-033": "2026-08-08",
    "TC-ADM-034": "2026-08-08",
    "TC-ADM-035": "2026-08-10",
    "TC-ADM-036": "2026-08-10",
    "TC-OPD-001": "2026-08-07",
    "TC-OPD-002": "2026-08-07",
    "TC-OPD-003": "2026-08-07",
    "TC-OPD-004": "2026-08-07",
    "TC-OPD-005": "2026-08-10",
    "TC-OPD-006": "2026-08-10",
    "TC-OPD-007": "2026-08-10",
    "TC-OPD-008": "2026-08-10",
    "TC-OPD-009": "2026-08-10",
    "TC-OPD-010": "2026-08-10",
    "TC-OPD-011": "2026-08-10",
    "TC-OPD-012": "2026-08-16",
    "TC-OPD-013": "2026-08-16",
    "TC-OPD-014": "2026-08-16",
    "TC-OPD-015": "2026-08-16",
    "TC-OPD-016": "2026-08-16",
    "TC-OPD-017": "2026-08-16",
    "TC-OPD-018": "2026-08-16",
}

# Notes stamped on rows added in later batches
NOTES_BY_ID = {
    "TC-ADM-022": "Added 2026-07-25 — Unknown Patient modal",
    "TC-ADM-023": "Added 2026-07-25 — Unknown Patient modal",
    "TC-ADM-024": "Added 2026-07-25 — Unknown Patient modal",
    "TC-ADM-025": "Added 2026-07-25 — Dynamic UI / default جهة",
    "TC-ADM-026": "Added 2026-07-25 — Dynamic label (إثبات الشخصية)",
    "TC-ADM-027": "Updated 2026-08-05 — نوع التعاقد default العضو نفسه + DDL options + درجة القرابة",
    "TC-ADM-028": "Added 2026-07-28 — Top action bar redirects",
    "TC-ADM-029": "Added 2026-07-28 — Barcode popup window",
    "TC-ADM-030": "Added 2026-07-28 — Under-dev action buttons",
    "TC-ADM-031": "Added 2026-07-28 — External visit family (same visit code)",
    "TC-ADM-032": "Added 2026-07-28 — Incompatible visit ERROR",
    "TC-ADM-033": "Added 2026-08-08 — Refresh patient profile → /reception",
    "TC-ADM-034": "Added 2026-08-08 — جديد on patient profile → /reception",
    "TC-ADM-035": "Added 2026-08-10 — Reservation Agenda modal current dates",
    "TC-ADM-036": "Added 2026-08-10 — Reservation Agenda search by specific date",
    "TC-OPD-001": "Added 2026-08-07 — OPD draft → services",
    "TC-OPD-002": "Added 2026-08-07 — Continue same visit (patient-data-continue-same-visit-btn)",
    "TC-OPD-003": "Added 2026-08-07 — Close unpaid visit blocked (Swal confirm + تنبيه)",
    "TC-OPD-004": "Added 2026-08-07 — Deferred (enabled=false) until clarified",
    "TC-OPD-005": "Added 2026-08-10 — OPD services page header & summary chips",
    "TC-OPD-006": "Added 2026-08-10 — OPD services page tabs & badge counts",
    "TC-OPD-007": "Added 2026-08-10 — Add service form initial state & price field restrictions",
    "TC-OPD-008": "Added 2026-08-10 — Add service full flow & save visit info",
    "TC-OPD-009": "Added 2026-08-10 — Save visit info button success toast",
    "TC-OPD-010": "Added 2026-08-10 — OPD payment flow & receipt history table",
    "TC-OPD-011": "Added 2026-08-10 — OPD services page back button navigation",
    "TC-OPD-012": "Added 2026-08-16 — Service group clinic filter",
    "TC-OPD-013": "Added 2026-08-16 — ICD10 mandatory check by entity",
    "TC-OPD-014": "Added 2026-08-16 — Print barcode button URL",
    "TC-OPD-015": "Added 2026-08-16 — Claim modal operations",
    "TC-OPD-016": "Added 2026-08-16 — Short code lookup (valid 4444 & invalid 91238128 error)",
    "TC-OPD-017": "Added 2026-08-16 — Approval number field state toggle (service 444)",
    "TC-OPD-018": "Added 2026-08-16 — Complete OPD service draft & save cycle",
}


def style_header(ws):
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:N1"


def main():
    wb = Workbook()

    # ---- Instructions sheet ----
    ws2 = wb.active
    ws2.title = "How to use"
    ws2["A1"] = "Paymax Automation — Test Case Register"
    ws2["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws2["A3"] = "Purpose / الغرض"
    ws2["A3"].font = Font(bold=True, size=12)
    ws2["A4"] = (
        "This workbook is the master list of all automated (and planned) test cases for Paymax."
    )
    ws2["A5"] = (
        "هذا الملف هو السجل الرئيسي لكل التست كيسز اللي اتعملت أو هتتعمل على نظام Paymax."
    )
    ws2["A7"] = "How to add a new test case / إضافة تيست كيس جديد"
    ws2["A7"].font = Font(bold=True, size=12)
    steps = [
        '1. Open the "Test Cases" sheet.',
        "2. Add a new row at the bottom.",
        "3. Fill TC ID using the next number (example: TC-ADM-022). Prefix: ADM = Admission / Medical File.",
        "4. Fill English + Arabic name and description.",
        "5. Choose Type / Priority / Automation Status from the dropdowns.",
        '6. Put the Java method name in "Method Name" when it is automated.',
        "",
        "1. افتح شيت Test Cases.",
        "2. أضف صف جديد في آخر الجدول.",
        "3. اكتب TC ID بالرقم التالي (مثال: TC-ADM-022). البادئة ADM = الملف الطبي.",
        "4. املأ الاسم والوصف بالعربي والإنجليزي.",
        "5. اختَر Type / Priority / Automation Status من القوائم.",
        "6. لما التيست يبقى Automated، حط اسم الميثود في Method Name.",
    ]
    for i, text in enumerate(steps):
        ws2.cell(8 + i, 1, text)

    automated = sum(1 for c in CASES if c[7] == "Automated")
    planned = sum(1 for c in CASES if c[7] == "Planned")

    ws2["A23"] = "Current coverage / التغطية الحالية"
    ws2["A23"].font = Font(bold=True, size=12)
    ws2["A24"] = "Modules: Admission (ADM) + Outpatient (OPD)"
    ws2["A25"] = f"Total test cases: {len(CASES)} | Automated: {automated} | Planned: {planned}"
    ws2["A26"] = f"Classes: {SUITE_ADMISSION} + {SUITE_ACTION_BAR} + {SUITE_OPD}"
    ws2["A27"] = f"Last updated: {TODAY}"
    ws2["A29"] = "Added on 2026-07-25 / اللي اتضاف يوم 25-07"
    ws2["A29"].font = Font(bold=True, size=12)
    ws2["A30"] = "TC-ADM-022 → Unknown patient: code is read-only"
    ws2["A31"] = "TC-ADM-023 → Unknown patient: happy path save (Success)"
    ws2["A32"] = "TC-ADM-024 → Unknown patient: close modal (X)"
    ws2["A33"] = "TC-ADM-025 → Default الجهة = $$نقدي 2019"
    ws2["A34"] = "TC-ADM-026 → إثبات الشخصية updates ID field label (e.g. جواز سفر)"
    ws2["A35"] = "TC-ADM-027 → نوع التعاقد default العضو نفسه; DDL: العضو نفسه|مريض تابع; مريض تابع → درجة القرابة الزوج/الزوجة"
    ws2["A37"] = "Added on 2026-07-28 / اللي اتضاف يوم 28-07 (Top Action Bar)"
    ws2["A37"].font = Font(bold=True, size=12)
    ws2["A38"] = "TC-ADM-028 → Action bar redirects with dynamic patientCode"
    ws2["A39"] = "TC-ADM-029 → Barcode popup (/print/barcode?id=) + close print window"
    ws2["A40"] = "TC-ADM-030 → Under-dev buttons: قسطرة / غسيل كلى / المناظير / الأرشيف"
    ws2["A41"] = "TC-ADM-031 → External visit family (OPD/Rays/Lab) keeps same كود الزيارة"
    ws2["A42"] = "TC-ADM-032 → Incompatible visit switch shows ERROR (Inpatient/Emergency vs others)"
    ws2["A44"] = "Added on 2026-08-07 / اللي اتضاف يوم 07-08 (Outpatient cycle)"
    ws2["A44"].font = Font(bold=True, size=12)
    ws2["A45"] = "TC-OPD-001 → Draft OPD visit save → /clinic/visits/{code}/services"
    ws2["A46"] = "TC-OPD-002 → Continue same visit (patient-data-continue-same-visit-btn)"
    ws2["A47"] = "TC-OPD-003 → Close unpaid visit blocked (confirm نعم → تنبيه لم تُدفع بعد)"
    ws2["A48"] = "TC-OPD-004 → Add service & total (Planned / enabled=false)"
    ws2["A50"] = "Added on 2026-08-08 / اللي اتضاف يوم 08-08 (Reception URL reset)"
    ws2["A50"].font = Font(bold=True, size=12)
    ws2["A51"] = "TC-ADM-033 → Refresh on /reception/patient/{code} redirects to /reception"
    ws2["A52"] = "TC-ADM-034 → جديد on selected patient navigates to /reception"
    ws2["A54"] = "Added on 2026-08-10 / اللي اتضاف يوم 10-08 (Reservation Agenda & OPD Services/Payments)"
    ws2["A54"].font = Font(bold=True, size=12)
    ws2["A55"] = "TC-ADM-035 → Reservation Agenda modal opens with current dates (from/to = today)"
    ws2["A56"] = "TC-ADM-036 → Reservation Agenda search by specific date (2026-08-08) returns matching results"
    ws2["A57"] = "TC-OPD-005 → OPD services page header title and summary chips (total, paid, insurance %, remaining, services count)"
    ws2["A58"] = "TC-OPD-006 → OPD services page tabs active state & badge counts"
    ws2["A59"] = "TC-OPD-007 → Add service form initial state & price field restrictions (readonly)"
    ws2["A60"] = "TC-OPD-008 → Add service to visit & save visit info"
    ws2["A61"] = "TC-OPD-009 → Save visit info button success toast notification"
    ws2["A62"] = "TC-OPD-010 → OPD visit payment flow (receipt created & added to history table)"
    ws2["A63"] = "TC-OPD-011 → Services page back button navigation"
    ws2["A65"] = "Seed patients (manual setup):"
    ws2["A65"].font = Font(bold=True, size=12)
    ws2["A66"] = "OPD active 30499 | Unpaid OPD 30552 | Inpatient 30496 | Emergency 30495 | Rays 30485 | Lab 30456"
    ws2["A68"] = "Suggested ID prefixes:"
    ws2["A68"].font = Font(bold=True, size=12)
    ws2["A69"] = "ADM = Patient Admission | OPD = Outpatient | BILL = Billing | LAB = Lab | PHARM = Pharmacy | AUTH = Login/Auth"
    ws2.column_dimensions["A"].width = 110

    # ---- Test Cases sheet ----
    ws = wb.create_sheet("Test Cases")
    style_header(ws)

    for i, case in enumerate(CASES, 1):
        tc_id, name_en, name_ar, desc_en, desc_ar, ttype, prio, status, method = case
        module_en, module_ar = MODULE_BY_ID.get(tc_id, (MODULE_ADM_EN, MODULE_ADM_AR))
        row = [
            tc_id,
            module_en,
            module_ar,
            name_en,
            name_ar,
            desc_en,
            desc_ar,
            ttype,
            prio,
            status,
            method,
            SUITE_BY_ID.get(tc_id, SUITE_ADMISSION),
            DATE_BY_ID.get(tc_id, TODAY),
            NOTES_BY_ID.get(tc_id, ""),
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(i + 1, col, val)
            cell.border = THIN
            cell.alignment = CENTER if col in (1, 8, 9, 10, 13) else WRAP
            if i % 2 == 0:
                cell.fill = ALT_FILL
        ws.row_dimensions[i + 1].height = 48

    widths = {
        "A": 14,
        "B": 38,
        "C": 34,
        "D": 55,
        "E": 50,
        "F": 55,
        "G": 50,
        "H": 12,
        "I": 10,
        "J": 16,
        "K": 42,
        "L": 24,
        "M": 12,
        "N": 30,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    dv_type = DataValidation(
        type="list",
        formula1='"Positive,Negative,Boundary,Smoke,Regression"',
        allow_blank=True,
    )
    dv_prio = DataValidation(
        type="list", formula1='"High,Medium,Low"', allow_blank=True
    )
    dv_status = DataValidation(
        type="list",
        formula1='"Automated,In Progress,Planned,Blocked,Manual Only"',
        allow_blank=True,
    )
    for dv, rng in (
        (dv_type, "H2:H500"),
        (dv_prio, "I2:I500"),
        (dv_status, "J2:J500"),
    ):
        dv.add(rng)
        ws.add_data_validation(dv)

    # ---- Summary sheet ----
    ws3 = wb.create_sheet("Summary")
    summary_headers = [
        "Module (EN)",
        "Module (AR)",
        "Total TCs",
        "Automated",
        "In Progress",
        "Planned",
    ]
    for col, header in enumerate(summary_headers, 1):
        cell = ws3.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
        cell.alignment = CENTER

    adm_cases = [c for c in CASES if not c[0].startswith("TC-OPD-")]
    opd_cases = [c for c in CASES if c[0].startswith("TC-OPD-")]

    def write_summary_row(row_idx, module_en, module_ar, cases):
        ws3.cell(row_idx, 1, module_en)
        ws3.cell(row_idx, 2, module_ar)
        ws3.cell(row_idx, 3, len(cases))
        ws3.cell(row_idx, 4, sum(1 for c in cases if c[7] == "Automated"))
        ws3.cell(row_idx, 5, sum(1 for c in cases if c[7] == "In Progress"))
        ws3.cell(row_idx, 6, sum(1 for c in cases if c[7] == "Planned"))
        for col in range(1, 7):
            ws3.cell(row_idx, col).border = THIN
            ws3.cell(row_idx, col).alignment = CENTER if col >= 3 else WRAP

    write_summary_row(2, "Patient Admission / Medical File", "الملف الطبي / استقبال المريض", adm_cases)
    write_summary_row(3, "Outpatient Visits (OPD)", "العيادات الخارجية", opd_cases)

    for col, width in zip("ABCDEF", [36, 32, 12, 12, 12, 12]):
        ws3.column_dimensions[col].width = width

    out_dated = r"d:\Ma3n\Automation Test For new Paymax\Paymax_Test_Cases 10-8-2026.xlsx"
    wb.save(OUT)
    wb.save(out_dated)
    print(f"Saved: {OUT}")
    print(f"Saved: {out_dated}")
    print(f"Total cases: {len(CASES)}")


if __name__ == "__main__":
    main()
